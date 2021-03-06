# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Alignment
import re

class MyParsing:
    def __init__(self):
        super().__init__()
        self.txtList = []
        self.wsList = []
        self.start = ""
        self.end = ""
        self.col = 0
        self.row = 0
        self.dir = ""

    def initVal(self, encode, start, end, col, row, width, height):
        self.isWindows = encode
        self.start = start
        self.end = end
        self.col = col
        self.row = row
        self.width = width
        self.height = height

    # FH는 파일핸들러
    # Return 값에 따른 에러처리 1:텍스트파일에러, 2:파싱문자열에러, 3:잘못된파싱, 4:인코딩에러
    def writeCell(self, filename):
        stcol = self.col
        strow = self.row

        wb = openpyxl.load_workbook(filename)
        FH = None
        # 워크시트 별로 작업 진행
        for wsname in self.txtList:
            selected_encode = ''
            if self.isWindows == 'True':
                selected_encode = 'cp949'
            else:
                selected_encode = 'utf-8'
            # 텍스트 파일을 연다. UTF-8로 인코딩 된 텍스트 파일만 불러올 수 있다.
            try:
                FH = open(self.dir+"/"+wsname, 'r', encoding=selected_encode)
            except IOError:
                # print("[-] 에러: 텍스트 파일 오픈 에러, 파일이 존재하는지 확인해주세요.")
                wb.close()
                return 1

            # 워크시트 선택, 세부설정
            # ws = wb.active active는 첫번째 워크시트를 선택하는 코드임
            ws = wb.get_sheet_by_name(os.path.splitext(wsname)[0])
            ws.column_dimensions[stcol].width = self.width

            # 정규표현식 컴파일
            pstart = re.compile(self.start)
            pend = re.compile(self.end)

            #텍스트 내용 가져오기
            contentList = []        #파싱포인트 배열
            nextcell = int(strow)   #셀 행 구분
            # 파일의 내용을 라인 단위로 리스트에 저장
            findPoint = False
            try:
                for line in FH:
                    if pstart.search(line) != None:
                        findPoint = True
                    #시작 포인트 찾았을때, 엔드포인트 진행
                    if findPoint:
                        contentList.append(line)
                         # End 문구를 찾았을 때, 리스트에 기록된 내용을 시트에 저장
                        if pend.search(line) != None:
                            # 첫번째, 마지막에 기록된 파싱문구를 지워서 최적화한다.
                            try:
                                contentList.pop(0)
                                contentList.pop(-1)
                            except IndexError:
                                #print("[-] 에러: 시작/끝 파싱포인트가 매칭되지 않았습니다. 문자열 또는 정규표현식을 다시 확인해주세요.")
                                wb.close()
                                FH.close()
                                return 2
                            # 마지막 개행 문자만 제거
                            contentList[-1] = contentList[-1].rstrip('\r')
                            contentList[-1] = contentList[-1].rstrip('\n')

                            # 셀에 데이터 기록
                            ws.row_dimensions[nextcell].height = self.height
                            ws[stcol + str(nextcell)].alignment = Alignment(vertical='top', wrap_text=True)
                            try:
                                ws[stcol + str(nextcell)] = "".join(contentList)
                            except openpyxl.utils.exceptions.IllegalCharacterError:
                                print("[셀 기록 불가] 한셀에 32700 문자를 넘어간 경우 또는 파싱문자열 시작-끝이 쌍으로 존재하지 않은 경우 발생(파싱불일치)")
                                pass
                                #return 3
                            # 리스트 초기화, 다음 엑셀 행 반복
                            contentList = []
                            findPoint = False
                            nextcell += 1
            except UnicodeError:
                #print("[-] 에러: 텍스트 파일의 인코딩 상태를 확인해주세요.")
                wb.close()
                FH.close()
                return 4

            # for문 바깥처리, 엑셀 저장 후 자원 반납
            wb.save(filename)
            wb.close()
            FH.close()

    #워크시트가 존재하지 않을 경우 새로 생성
    def checkSheet(self, filename):
        wb = openpyxl.load_workbook(filename)
        # wb.get_sheet_names() 모든 워크시트 이름 가져오기
        for wsname in self.wsList:
            try:
                ws = wb[wsname]
            except KeyError:
                ws = wb.create_sheet()
                ws.title = wsname
            finally:
                try:
                    wb.save(filename)
                except PermissionError:
                    wb.close()
                    return 1
        wb.close()

    # 현재 디렉토리 파일 목록 획득, 텍스트 파일만 추출하기
    def getTxt(self, dir):
        self.dir = dir
        fileList = os.listdir(dir)
        for file in fileList:
            if file.find('.txt') > 0:
                self.txtList.append(file)
                self.wsList.append(os.path.splitext(file)[0]) # 확장자 제거